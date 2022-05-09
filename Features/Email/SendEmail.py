import openpyxl
import smtplib
from email.message import EmailMessage
import os
from Features.Email.AddEmail import addEmail
from Functions.TakeCommand import takeCommand
from Functions.Speak import speak


def sendEmail(receiver, subject, message):
    try:
        server = smtplib.SMTP('smtp.gmail.com', 587)
        server.starttls()
        server.login('asatiramanuj@gmail.com', 'RAmanUJ$#@!123')
        email = EmailMessage()
        email['From'] = 'asatiramanuj@gmail.com'
        email['To'] = receiver
        email['Subject'] = subject
        email.set_content(message)
        server.send_message(email)
        speak("Email send successfully")
        return
    except:
        speak("Sorry, I can't send the mail some error on the server side")
        return

email_list = {}

def getEmailList():
    path = os.path.realpath(os.path.join(os.path.dirname(__file__), '..', '..', 'Database', 'EmailDB', 'emailDB.xlsx'))

    wb_obj = openpyxl.load_workbook(path)

    sheet_obj = wb_obj.active
    m_row = sheet_obj.max_row

    for i in range(1, m_row + 1):
        cell_obj1 = sheet_obj.cell(row=i, column=1)
        cell_obj2 = sheet_obj.cell(row=i, column=2)
        email_list.__setitem__(cell_obj1.value, cell_obj2.value)


def get_email_info():
    getEmailList()
    speak('To Whom you want to send email')
    name = takeCommand()

    if name in email_list:
        pass
    else:
        speak(f"{name} is not available in the database")
        speak(f"Would you like to add {name} in the database")
        choice = takeCommand()
        # choice = input()
        if choice == 'yes':
            speak("Okey")
            addEmail()
            get_email_info()
        elif choice == 'no':
            speak("Okey")
            return
        else:
            speak("Pardon, I can't understand")
            return

    receiver = email_list[name]
    print(receiver)
    speak('Tell me the subject of your email')
    subject = takeCommand()
    speak('Tell me the message in your email')
    message = takeCommand()
    sendEmail(receiver, subject, message)
    return
